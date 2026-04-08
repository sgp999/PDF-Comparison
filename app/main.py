from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.responses import FileResponse, HTMLResponse
from pypdf import PdfReader
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openai import OpenAI
from dotenv import load_dotenv
import os
import uuid
import json
import re

load_dotenv()

client = OpenAI(api_key=os.getenv("OPENAI_API_KEY"))
app = FastAPI()


def extract_text_from_pdf(upload_file: UploadFile) -> str:
    try:
        reader = PdfReader(upload_file.file)
        text = ""

        for page in reader.pages:
            page_text = page.extract_text()
            if page_text:
                text += page_text + "\n"

        return text.strip()

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error reading PDF: {str(e)}")


def clean_json_response(raw_text: str) -> str:
    raw_text = raw_text.strip()

    if raw_text.startswith("```"):
        raw_text = re.sub(r"^```json", "", raw_text)
        raw_text = re.sub(r"^```", "", raw_text)
        raw_text = re.sub(r"```$", "", raw_text)
        raw_text = raw_text.strip()

    return raw_text


def extract_insurance_data(text: str) -> dict:
    prompt = f"""
You are an expert at reading insurance policy documents.

Extract structured data from the text below.

Return ONLY valid JSON.
Do not include markdown.
Do not include explanation.

Rules:
- If a value is not found, return "N/A"
- Keep dollar amounts exactly as written
- Do not guess
- If multiple values appear, choose the clearest standard benefit amount

Fields:
- plan_name
- premium
- deductible_individual
- deductible_family
- out_of_pocket_max_individual
- out_of_pocket_max_family
- primary_care
- specialist
- urgent_care
- emergency_room
- hospital_stay
- outpatient_surgery
- diagnostic_test
- imaging
- lab_work
- preventive_care
- mental_health_outpatient
- mental_health_inpatient
- physical_therapy
- ambulance
- generic_drugs
- preferred_brand_drugs

Policy text:
{text[:12000]}
"""

    try:
        response = client.responses.create(
            model="gpt-4o-mini",
            input=prompt
        )

        raw = response.output[0].content[0].text
        cleaned = clean_json_response(raw)
        data = json.loads(cleaned)

        if isinstance(data, dict):
            return data

        return {}

    except Exception:
        return {}


def build_comparison_rows(data1: dict, data2: dict):
    fields = [
        ("Plan Name", "plan_name"),
        ("Premium", "premium"),
        ("Deductible (Individual)", "deductible_individual"),
        ("Deductible (Family)", "deductible_family"),
        ("Out of Pocket Max (Individual)", "out_of_pocket_max_individual"),
        ("Out of Pocket Max (Family)", "out_of_pocket_max_family"),
        ("Primary Care", "primary_care"),
        ("Specialist", "specialist"),
        ("Urgent Care", "urgent_care"),
        ("Emergency Room", "emergency_room"),
        ("Hospital Stay", "hospital_stay"),
        ("Outpatient Surgery", "outpatient_surgery"),
        ("Diagnostic Test", "diagnostic_test"),
        ("Imaging", "imaging"),
        ("Lab Work", "lab_work"),
        ("Preventive Care", "preventive_care"),
        ("Mental Health (Outpatient)", "mental_health_outpatient"),
        ("Mental Health (Inpatient)", "mental_health_inpatient"),
        ("Physical Therapy", "physical_therapy"),
        ("Ambulance", "ambulance"),
        ("Generic Drugs", "generic_drugs"),
        ("Preferred Brand Drugs", "preferred_brand_drugs"),
    ]

    rows = []

    for label, key in fields:
        rows.append({
            "Field": label,
            "Policy 1": data1.get(key, "N/A"),
            "Policy 2": data2.get(key, "N/A")
        })

    return rows


def safe_filename(name: str) -> str:
    name = name.replace(".pdf", "")
    name = re.sub(r"[^A-Za-z0-9_-]+", "_", name)
    return name[:40]


def save_to_excel(rows, output_path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Comparison Results"

    ws.append(["Field", "Policy 1", "Policy 2"])

    header_fill = PatternFill(fill_type="solid", start_color="D9EAF7", end_color="D9EAF7")
    bold_font = Font(bold=True)

    for cell in ws[1]:
        cell.font = bold_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row in rows:
        ws.append([
            row.get("Field", ""),
            row.get("Policy 1", ""),
            row.get("Policy 2", "")
        ])

    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 38
    ws.freeze_panes = "A2"

    wb.save(output_path)


@app.get("/")
def root():
    return {"message": "API is running"}


@app.get("/health")
def health():
    return {"status": "ok"}


@app.get("/ui", response_class=HTMLResponse)
def ui():
    return """
    <html>
        <head>
            <title>Insurance PDF Comparison</title>
            <style>
                body {
                    font-family: Arial, sans-serif;
                    max-width: 900px;
                    margin: 40px auto;
                    padding: 20px;
                    background: #f7f9fc;
                }
                .card {
                    background: white;
                    padding: 30px;
                    border-radius: 12px;
                    box-shadow: 0 2px 10px rgba(0,0,0,0.08);
                }
                h1 {
                    margin-top: 0;
                }
                .button {
                    background: #2563eb;
                    color: white;
                    border: none;
                    padding: 12px 18px;
                    border-radius: 8px;
                    cursor: pointer;
                    font-size: 16px;
                }
                .button:hover {
                    background: #1d4ed8;
                }
                input[type="file"] {
                    margin-top: 8px;
                    margin-bottom: 20px;
                }
                .note {
                    color: #555;
                    font-size: 14px;
                }
            </style>
        </head>
        <body>
            <div class="card">
                <h1>Compare Insurance PDFs</h1>
                <p class="note">Upload two policy PDFs to compare benefits and costs side by side.</p>

                <form action="/compare" enctype="multipart/form-data" method="post" onsubmit="showLoading()">
                    <label><strong>Policy 1 PDF</strong></label><br>
                    <input type="file" name="file1" accept=".pdf" required><br>

                    <label><strong>Policy 2 PDF</strong></label><br>
                    <input type="file" name="file2" accept=".pdf" required><br>

                    <button class="button" type="submit">Compare Files</button>
                </form>

                <p id="loading" style="display:none; margin-top:20px;"><strong>Processing PDFs...</strong></p>
            </div>

            <script>
                function showLoading() {
                    document.getElementById("loading").style.display = "block";
                }
            </script>
        </body>
    </html>
    """


@app.post("/compare", response_class=HTMLResponse)
async def compare_pdfs(file1: UploadFile = File(...), file2: UploadFile = File(...)):
    if not file1.filename.lower().endswith(".pdf") or not file2.filename.lower().endswith(".pdf"):
        raise HTTPException(status_code=400, detail="Both files must be PDFs.")

    text1 = extract_text_from_pdf(file1)
    file1.file.seek(0)

    text2 = extract_text_from_pdf(file2)
    file2.file.seek(0)

    data1 = extract_insurance_data(text1)
    data2 = extract_insurance_data(text2)

    comparison_rows = build_comparison_rows(data1, data2)

    os.makedirs("exports", exist_ok=True)

    file_id = str(uuid.uuid4())

    file1_base = safe_filename(file1.filename)
    file2_base = safe_filename(file2.filename)
    download_name = f"{file1_base}_vs_{file2_base}.xlsx"
    output_path = os.path.join("exports", f"{file_id}.xlsx")

    save_to_excel(comparison_rows, output_path)

    table_rows = ""
    for row in comparison_rows:
        table_rows += f"""
        <tr>
            <td>{row.get("Field", "")}</td>
            <td>{row.get("Policy 1", "")}</td>
            <td>{row.get("Policy 2", "")}</td>
        </tr>
        """

    return f"""
    <html>
        <head>
            <title>Comparison Results</title>
            <style>
                body {{
                    font-family: Arial, sans-serif;
                    max-width: 1100px;
                    margin: 40px auto;
                    padding: 20px;
                    background: #f7f9fc;
                }}
                .card {{
                    background: white;
                    padding: 30px;
                    border-radius: 12px;
                    box-shadow: 0 2px 10px rgba(0,0,0,0.08);
                }}
                table {{
                    width: 100%;
                    border-collapse: collapse;
                    margin-top: 20px;
                }}
                th, td {{
                    border: 1px solid #dcdcdc;
                    padding: 10px;
                    text-align: left;
                    vertical-align: top;
                }}
                th {{
                    background: #e8f0fe;
                }}
                .button {{
                    display: inline-block;
                    background: #2563eb;
                    color: white;
                    text-decoration: none;
                    padding: 12px 18px;
                    border-radius: 8px;
                    margin-top: 20px;
                }}
                .button:hover {{
                    background: #1d4ed8;
                }}
                .secondary {{
                    background: #64748b;
                    margin-left: 10px;
                }}
            </style>
        </head>
        <body>
            <div class="card">
                <h1>Comparison Results</h1>

                <table>
                    <tr>
                        <th>Field</th>
                        <th>Policy 1</th>
                        <th>Policy 2</th>
                    </tr>
                    {table_rows}
                </table>

                <a class="button" href="/download/{file_id}?name={download_name}">Save to Excel</a>
                <a class="button secondary" href="/ui">Compare Another Pair</a>
            </div>
        </body>
    </html>
    """


@app.get("/download/{file_id}")
def download_file(file_id: str, name: str = "comparison_results.xlsx"):
    file_path = os.path.join("exports", f"{file_id}.xlsx")

    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail="File not found.")

    return FileResponse(
        path=file_path,
        filename=name,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )